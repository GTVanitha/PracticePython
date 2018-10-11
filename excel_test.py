import openpyxl

wb = openpyxl.load_workbook(filename = 'som_programs/sample_excel.xlsx')

sheets = wb.sheetnames
print sheets

ws = wb[sheets[1]]


m = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'I', 'J']
# find the len of sheet / row length
# take 

col_len = len(ws[1])

print "column length >:", col_len

#for a in ws:
#    print a

cols = m[:col_len]

print "ols: ", cols

#print "-========="
counter = 0
columns = []
for c in cols:
    columns.insert(counter, ws[c])
    counter += 1


values_lis = []
for col_ref in columns:
    col_values = []
    for col_cell in col_ref:
        col_values.append(col_cell.value)
    values_lis.append(col_values)

print values_lis
print "=======end===="

        

def myfunc(a):
    values = []
    for v in a:
          values.append(v.value)
    return values

x = map(myfunc, columns)

#print x


names = x[0] 
age = x[1]
dob = x[2]

inp = raw_input("Enter a name to figure out the deatils of the person:");
c = 0
nothing = 1
for n in names:
    if n == inp:
        nothing = 0
        print "Name:%s, Age: %d, DOB: %s" %(names[c], age[c], dob[c])
    c +=1

if nothing == 1:
    print "Not able find the record of %s" %(inp)

print "=====ENNNDDDD-----"



'''


print "===="

f_col = ws['A']
s_col= ws['B']

print s_col[1].value

c = 0
names = []

for c1 in f_col:
    if c == 0:
        c += 1
        continue 
    names.append(c1.value)

print names
 
c = 0
age = []

for c2 in s_col:
    if c == 0:
        c+= 1
        continue
    age.append(c2.value)

print age

d = dict(zip(names, age))

print d
'''
